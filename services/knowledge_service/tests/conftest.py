import pytest
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage


@pytest.fixture
def sample_xlsx_path(tmp_path):
    """1 sheet, vài dòng dữ liệu, 1 ảnh nhúng neo tại cell C2 (row 2)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "ID"
    ws["B1"] = "Mo_ta"
    ws["A2"] = "REQ-1"
    ws["B2"] = "Doi mau nut dang nhap"
    ws["A3"] = "REQ-2"
    ws["B3"] = "Them truong ghi chu"

    tiny_png_path = tmp_path / "tiny.png"
    PILImage.new("RGB", (4, 4), color=(255, 0, 0)).save(tiny_png_path)
    ws.add_image(XLImage(str(tiny_png_path)), "C2")

    xlsx_path = tmp_path / "sample.xlsx"
    wb.save(xlsx_path)
    return str(xlsx_path)
